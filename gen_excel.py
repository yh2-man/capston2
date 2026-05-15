import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb = openpyxl.Workbook()

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="2E75B6")
cat_fill = PatternFill("solid", fgColor="D6E4F0")
warn_fill = PatternFill("solid", fgColor="FFF2CC")
err_fill = PatternFill("solid", fgColor="FCE4EC")
ok_fill = PatternFill("solid", fgColor="E8F5E9")
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
wrap = Alignment(wrap_text=True, vertical='top')

def style_header(ws, cols):
    for i,c in enumerate(cols,1):
        cell = ws.cell(1,i,c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = wrap; cell.border = thin

def add_row(ws, r, vals, fill=None):
    for i,v in enumerate(vals,1):
        cell = ws.cell(r,i,v)
        cell.alignment = wrap; cell.border = thin
        if fill: cell.fill = fill

# ===== Sheet 1: Frontend =====
ws1 = wb.active; ws1.title = "프론트엔드(Flutter)"
cols1 = ["기능영역","파일명","클래스/함수","역할","호출하는 백엔드 API","사용여부","비고"]
style_header(ws1, cols1)
fe_data = [
 ["인증","main.dart","main(), _isTokenExpired, _initFcm, _requestLocationPermission, OurTApp","앱 진입점, 토큰검증, FCM초기화, 위치권한","없음(로컬)","O",""],
 ["인증","login_screen.dart","LoginScreen, _loginWithKakao, _buildLogo, _buildKakaoButton","카카오 OAuth 로그인 UI","POST /api/v1/auth/oauth/kakao","O",""],
 ["인증","profile_setup_screen.dart","ProfileSetupScreen, _submitProfile, _pickProfileImage","신규유저 프로필설정","POST /api/v1/auth/profile-setup, POST /api/v1/auth/fcm-token","O","이미지업로드 TODO"],
 ["인증","auth_api_service.dart","AuthApiService: getKakaoAuthUrl, sendAuthCode, saveFcmToken, setupProfile","인증 API 호출","POST /api/v1/auth/oauth/{provider}, POST /api/v1/auth/fcm-token, POST /api/v1/auth/profile-setup","O",""],
 ["설정","api_config.dart","ApiConfig: kakaoMapKey, nongsaroKey, tourApiKey, backendBaseUrl, mapApiBaseUrl","API키/URL 중앙관리","없음","O",""],
 ["챗봇","chatbot_service.dart","ChatbotService: sendMessage, streamMessage, clearSession + 응답DTO들","챗봇 API 호출(일반+SSE)","POST /chatbot/message, POST /chatbot/message/stream, DELETE /chatbot/session/{id}","O",""],
 ["챗봇","chat_screen.dart","ChatScreen: _sendMessage, _resetSession, _buildAgentActivityPanel 등","전체화면 챗봇 UI","(chatbot_service 경유)","O",""],
 ["챗봇","chat_floating_button.dart","ChatFloatingButton: _sendMessage, _buildHistoryPanel, _buildInputDock","플로팅 챗봇 위젯","(chatbot_service 경유)","O",""],
 ["챗봇","app_action_runtime.dart","AppActionRuntime: execute, _isMapAction, _isScreenAction, _screenFor","챗봇 액션→화면이동","없음(클라이언트)","O",""],
 ["홈","main_screen.dart","MainScreen: _loadPosts, _loadFestivals, _sendMainChatMessage, _buildShortcutButtons 등","메인 홈 화면","(community_api, tour_api, chatbot 경유)","O","764줄, 리팩토링 권장"],
 ["지도","kakao_map_screen.dart","KakaoMapScreen: setSearchQuery, zoomIn/Out, moveToCurrentLocation, _loadFestivalData 등","카카오맵+축제표시","(tour_api 경유)","O","977줄, 리팩토링 권장"],
 ["커뮤니티","community_feed_screen.dart","CommunityFeedScreen: _loadPosts, _toggleLike, _toggleSave, _openCreatePost","커뮤니티 피드","(community_api 경유)","O",""],
 ["커뮤니티","community_api_service.dart","CommunityApiService: getPosts, createPost, toggleLike, toggleSave + CommunityPost DTO","커뮤니티 API 호출","GET/POST /api/v1/community/posts, POST .../like, POST .../save","O",""],
 ["커뮤니티","create_post_screen.dart","CreatePostScreen: _pickImage, _submit","일반 게시글 작성","(community_api 경유)","X 미사용","어디서도 import 안됨"],
 ["꽃스팟","create_flower_spot_screen.dart","CreateFlowerSpotScreen: _takePicture, _identifyPlant, _submit, _toggleLocationShare","꽃 사진+위치 게시","(flower_spot_api 경유)","O",""],
 ["꽃스팟","flower_spot_api_service.dart","FlowerSpotApiService: identifyPlant, createFlowerSpot, getFlowerSpots","꽃스팟 API 호출","POST /api/v1/flower-spots/identify, POST/GET /api/v1/flower-spots","O",""],
 ["꽃도감","flower_book_page.dart","FlowerBookPage: _loadFlowers, _search, _showFlowerDetail + _FlowerDetailSheet","월별 꽃 도감","(flower_book_api 경유)","O",""],
 ["꽃도감","flower_book_api_service.dart","FlowerBookApiService: getByMonth, getDetail, search","꽃 도감 API 호출","GET /api/v1/flowers/monthly/{m}, GET /api/v1/flowers/{id}, GET /api/v1/flowers/search","O",""],
 ["꽃API","flower_api_service.dart","FlowerApiService: getFlowerList, getFlowerDetail, getFlowerByDate + FlowerData, FlowerDetail","공공데이터 오늘의꽃 API","외부: apis.data.go.kr","X 미사용","어디서도 import 안됨"],
 ["축제","tour_api_service.dart","TourApiService: getFlowerFestivals, searchFlowerFestivals + FestivalData","한국관광공사 축제API","외부: apis.data.go.kr TourAPI","O",""],
 ["산책","pedometer_screen.dart","PedometerScreen: _loadData, _simulateWalk, _buildCircularProgress, _buildWeeklyChart","만보기+포인트UI","(walk_api 경유)","O",""],
 ["산책","walk_api_service.dart","WalkApiService: getWeeklyRecords, getPointBalance, syncSteps","산책 API 호출","GET /api/v1/walk/records/weekly, GET /api/v1/walk/points, POST /api/v1/walk/sync","O","백엔드 미구현"],
 ["저장","saved_page.dart","SavedPage: _loadData, _buildSavedPosts, _buildSavedSpots","저장된 게시글/스팟","(saved_api 경유)","O",""],
 ["저장","saved_api_service.dart","SavedApiService: getSavedPosts, getSavedSpots, unsavePost, unsaveSpot","저장 API 호출","GET/DELETE /api/v1/saved/posts, GET/DELETE /api/v1/saved/spots","O","백엔드 미구현"],
 ["내정보","my_info_screen.dart","MyInfoScreen: _loadUserInfo","내 정보 화면","없음(로컬)","O","최소구현"],
 ["테마","season_theme.dart","SeasonTheme, SeasonColors","계절별 테마색상","없음","O",""],
 ["테마","app_colors.dart, app_dimensions.dart, app_text_styles.dart, app_theme.dart","디자인 토큰","색상/크기/텍스트 상수","없음","△ 부분사용","SeasonTheme과 중복 가능"],
 ["위젯","app_bottom_navigation.dart","AppBottomNavigation, AppNavTab","하단 네비게이션바","없음","O",""],
 ["위젯","message_bubble.dart","MessageBubble","챗봇 말풍선 위젯","없음","O",""],
 ["위젯","map_html_view*.dart","MapHtmlView (web/stub)","웹용 지도 뷰","없음","O",""],
 ["모델","chat_action.dart, chat_message.dart","ChatAction, ChatMessage, MessageRole","채팅 데이터 모델","없음","O",""],
 ["스텁","chatbot_screen.dart","ChatbotScreen, CommunityScreen, WalkScreen, FlowerBookScreen, SavedScreen","래퍼 위젯 모음","없음","△","직접 사용처 없음"],
 ["스텁","community_screen.dart, flower_book_screen.dart, saved_screen.dart, walk_screen.dart","export 'chatbot_screen.dart'","잘못된 re-export","없음","X 미사용","삭제 권장"],
]
for i,row in enumerate(fe_data,2):
    f = None
    if "미사용" in row[5]: f = err_fill
    elif "△" in row[5]: f = warn_fill
    add_row(ws1, i, row, f)
for col in ws1.columns:
    ws1.column_dimensions[col[0].column_letter].width = 22

# ===== Sheet 2: Backend =====
ws2 = wb.create_sheet("백엔드(Spring)")
cols2 = ["기능영역","파일명","클래스/메서드","역할","API 엔드포인트","프론트 연동","비고"]
style_header(ws2, cols2)
be_data = [
 ["인증","AuthController.java","refresh, setupProfile, oauthKakao, saveFcmToken, logout","인증 REST 컨트롤러","POST /api/v1/auth/refresh, /profile-setup, /oauth/kakao, /fcm-token, /logout","O",""],
 ["인증","AuthService.java","processOAuth, setupProfile, refresh, saveFcmToken, logout, buildLoginResponse","인증 비즈니스로직","(Controller 경유)","O",""],
 ["인증","OAuthService.java","processKakao, getKakaoAccessToken, getKakaoUserInfo","카카오 OAuth 처리","외부: kauth.kakao.com, kapi.kakao.com","O",""],
 ["인증","JwtProvider.java","generateAccessToken, generateRefreshToken, generateTempToken, validateToken, getUserId 등","JWT 토큰 유틸","없음","O",""],
 ["인증","JwtAuthenticationFilter.java","doFilterInternal, resolveToken","JWT 인증 필터","없음","O",""],
 ["인증","SecurityConfig.java","filterChain, corsConfigurationSource, passwordEncoder","보안 설정","없음","O","passwordEncoder 미사용"],
 ["인증","User.java","User 엔티티","사용자 DB 모델","없음","O",""],
 ["인증","UserRepository.java","findByProviderAndProviderId, existsByNickname 등","사용자 JPA 리포지토리","없음","O",""],
 ["인증","AuthDto.java","LoginResponse, RefreshRequest/Response, ProfileSetupRequest, OAuthRequest 등","인증 DTO","없음","O",""],
 ["인증","OAuthProperties.java","kakao.clientId, clientSecret","OAuth 설정값","없음","O",""],
 ["인증","AuthException.java","AuthException","인증 예외","없음","O",""],
 ["인증","OAuthCallbackController.java","OAuth 콜백","리다이렉트 처리","GET /oauth/callback","O",""],
 ["챗봇","ChatbotController.java","sendMessage, streamMessage, clearSession","챗봇 REST+SSE","POST /chatbot/message, /message/stream, DELETE /chatbot/session/{id}","O",""],
 ["챗봇","ChatbotService.java","chat, chatStream, clearSession + Agent 로직","AI 챗봇 핵심 서비스","없음","O","36KB, 대형파일"],
 ["챗봇","ChatActionValidator.java","액션 검증","챗봇 액션 유효성","없음","O",""],
 ["챗봇","ChatbotActionContext.java","액션 컨텍스트","Tool 실행 컨텍스트","없음","O",""],
 ["챗봇","DTO: ChatMessageRequest/Response, ChatAction, RoutePlan, AgentRunTrace, ToolResult","챗봇 DTO들","요청/응답 모델","없음","O",""],
 ["챗봇","ChatbotConfig.java","설정","챗봇 설정","없음","O",""],
 ["커뮤니티","CommunityController.java","getFeed, createPost, toggleLike, toggleSave","커뮤니티 REST","GET/POST /api/v1/community/posts, .../like, .../save","O",""],
 ["커뮤니티","FlowerSpotController.java","identify, createFlowerSpot, getFlowerSpots","꽃스팟 REST","POST /api/v1/flower-spots/identify, POST/GET /api/v1/flower-spots","O",""],
 ["커뮤니티","CommunityService.java","getFeed, createPost, toggleLike, toggleSave, createFlowerSpot, getFlowerSpots","커뮤니티 비즈니스로직","(Controller 경유)","O",""],
 ["커뮤니티","CommunityPost.java + CommunityDto.java","엔티티+DTO","게시글 모델","없음","O",""],
 ["커뮤니티","PostLike/SavedPost + Repository","좋아요/저장 엔티티","관계 테이블","없음","O",""],
 ["꽃도감","FlowerController.java","getCategories, getByCategory, getByMonth, getDetail, search, match","꽃 REST","GET /api/v1/flowers/**","O",""],
 ["꽃도감","FlowerService.java","getCategories, getFlowersByCategory/Month, getFlowerDetail, search, matchByScientificName","꽃 비즈니스로직","(Controller 경유)","O",""],
 ["꽃도감","FlowerBook/Category/SpeciesMapping + Repository","엔티티+리포지토리","꽃 DB 모델","없음","O",""],
 ["꽃도감","FlowerDto.java","CategoryResponse, FlowerSummary, FlowerDetail, SearchResult, MatchResult","꽃 DTO","없음","O",""],
 ["꽃도감","FlowerImportController.java","관리자 꽃 데이터 임포트","어드민 API","POST /api/v1/admin/flowers/import 등","X 관리용",""],
 ["꽃도감","NongsaroImportService.java, NongsaroImageService.java, WikiImageService.java","농사로/위키 이미지 크롤러","데이터 수집","외부 API","X 관리용",""],
 ["식물인식","PlantIdService.java","identify, parseResponse, findKoreanName","Plant.id API 호출","외부: api.plant.id","O",""],
 ["스토리지","StorageService.java (interface)","upload","파일 업로드 인터페이스","없음","O",""],
 ["스토리지","LocalStorageService.java","로컬 파일 저장","개발용 로컬 저장","없음","O",""],
 ["스토리지","OracleStorageService.java","OCI Object Storage","운영용 클라우드 저장","외부: OCI","O",""],
 ["공통","ApiResponse, dto, exception","응답 래퍼, 예외 처리","공통 유틸","없음","O",""],
 ["설정","AppConfig.java","RestTemplate Bean","HTTP 클라이언트 설정","없음","O",""],
 ["설정","AuditingConfig.java","JPA Auditing","생성/수정일 자동기록","없음","O",""],
]
for i,row in enumerate(be_data,2):
    f = None
    if "미사용" in str(row[5]) or "관리용" in str(row[5]): f = warn_fill
    add_row(ws2, i, row, f)
for col in ws2.columns:
    ws2.column_dimensions[col[0].column_letter].width = 22

# ===== Sheet 3: API 연관관계 =====
ws3 = wb.create_sheet("프론트↔백엔드 연관")
cols3 = ["API 엔드포인트","HTTP","프론트 호출 파일","백엔드 처리 파일","연동상태","비고"]
style_header(ws3, cols3)
api_data = [
 ["POST /api/v1/auth/oauth/kakao","POST","auth_api_service.dart → login_screen.dart","AuthController → OAuthService","O 정상",""],
 ["POST /api/v1/auth/profile-setup","POST","auth_api_service.dart → profile_setup_screen.dart","AuthController → AuthService","O 정상",""],
 ["POST /api/v1/auth/fcm-token","POST","auth_api_service.dart → login/profile","AuthController → AuthService","O 정상",""],
 ["POST /api/v1/auth/refresh","POST","미구현(프론트)","AuthController → AuthService","△ 프론트 미연동","토큰갱신 로직 필요"],
 ["POST /api/v1/auth/logout","POST","미구현(프론트)","AuthController → AuthService","△ 프론트 미연동","로그아웃 버튼 필요"],
 ["POST /chatbot/message","POST","chatbot_service.dart","ChatbotController → ChatbotService","O 정상",""],
 ["POST /chatbot/message/stream","POST","chatbot_service.dart (SSE)","ChatbotController → ChatbotService","O 정상",""],
 ["DELETE /chatbot/session/{id}","DELETE","chatbot_service.dart","ChatbotController → ChatbotService","O 정상",""],
 ["GET /api/v1/community/posts","GET","community_api_service.dart","CommunityController → CommunityService","O 정상",""],
 ["POST /api/v1/community/posts","POST","community_api_service.dart","CommunityController → CommunityService","O 정상",""],
 ["POST .../posts/{id}/like","POST","community_api_service.dart","CommunityController → CommunityService","O 정상",""],
 ["POST .../posts/{id}/save","POST","community_api_service.dart","CommunityController → CommunityService","O 정상",""],
 ["POST /api/v1/flower-spots/identify","POST","flower_spot_api_service.dart","FlowerSpotController → PlantIdService","O 정상",""],
 ["POST /api/v1/flower-spots","POST","flower_spot_api_service.dart","FlowerSpotController → CommunityService","O 정상",""],
 ["GET /api/v1/flower-spots","GET","flower_spot_api_service.dart","FlowerSpotController → CommunityService","O 정상",""],
 ["GET /api/v1/flowers/monthly/{m}","GET","flower_book_api_service.dart","FlowerController → FlowerService","O 정상",""],
 ["GET /api/v1/flowers/{id}","GET","flower_book_api_service.dart","FlowerController → FlowerService","O 정상",""],
 ["GET /api/v1/flowers/search","GET","flower_book_api_service.dart","FlowerController → FlowerService","O 정상",""],
 ["GET /api/v1/walk/**","GET/POST","walk_api_service.dart","없음 (백엔드 미구현)","X 백엔드 없음","Controller/Service 구현 필요"],
 ["GET/DELETE /api/v1/saved/**","GET/DELETE","saved_api_service.dart","없음 (백엔드 미구현)","X 백엔드 없음","Controller/Service 구현 필요"],
]
for i,row in enumerate(api_data,2):
    f = None
    if "X" in row[4]: f = err_fill
    elif "△" in row[4]: f = warn_fill
    else: f = ok_fill
    add_row(ws3, i, row, f)
for col in ws3.columns:
    ws3.column_dimensions[col[0].column_letter].width = 28

# ===== Sheet 4: 최적화 제안 =====
ws4 = wb.create_sheet("최적화 및 UI개선")
cols4 = ["구분","항목","현재상태","개선방안","우선순위"]
style_header(ws4, cols4)
opt_data = [
 ["코드정리","create_post_screen.dart","어디서도 import 안됨, 미사용","삭제 (create_flower_spot_screen이 대체)","높음"],
 ["코드정리","flower_api_service.dart","어디서도 import 안됨, 미사용","삭제 (flower_book_api_service가 대체)","높음"],
 ["코드정리","community_screen/flower_book_screen/saved_screen/walk_screen.dart","chatbot_screen.dart를 잘못 re-export","전부 삭제","높음"],
 ["코드정리","SecurityConfig.passwordEncoder()","소셜전용이라 BCrypt 미사용","Bean 제거","낮음"],
 ["리팩토링","main_screen.dart (764줄)","홈화면에 챗봇+축제+게시글+산책 모두 포함","위젯 분리: FestivalSection, PostPreviewStrip, WalkSummary 등","높음"],
 ["리팩토링","kakao_map_screen.dart (977줄)","지도+축제+검색+UI 한 파일","MapController 로직분리, FestivalOverlay 위젯 분리","높음"],
 ["리팩토링","chatbot_service.dart 내 DTO들","서비스파일에 6개 클래스 혼재","models/ 폴더로 DTO 분리","중간"],
 ["기능보완","토큰 자동갱신","refresh API 있지만 프론트 미연동","Dio Interceptor로 401시 자동 refresh 구현","높음"],
 ["기능보완","로그아웃","백엔드 logout 있지만 프론트 버튼 없음","MyInfoScreen에 로그아웃 버튼 추가","높음"],
 ["기능보완","Walk API 백엔드","프론트에서 호출하지만 백엔드 미구현","WalkController, WalkService, WalkRecord 엔티티 구현","높음"],
 ["기능보완","Saved API 백엔드","프론트에서 호출하지만 백엔드 미구현","SavedController, SavedService 구현","높음"],
 ["기능보완","프로필 이미지 업로드","profile_setup_screen에 TODO 표시","image_picker + StorageService 연동","중간"],
 ["UI개선","홈 산책요약","하드코딩 텍스트만 표시","WalkApiService 연동하여 실제 데이터 표시","중간"],
 ["UI개선","내정보 화면","닉네임+프로필만 표시, 최소구현","설정, 알림관리, 로그아웃, 내 게시글 등 추가","중간"],
 ["UI개선","에러 처리 통일","각 화면마다 다른 에러 표시 방식","공통 ErrorWidget, LoadingWidget 컴포넌트화","낮음"],
 ["UI개선","커뮤니티 댓글","댓글 아이콘만 있고 기능 없음","댓글 기능 구현 (Comment 엔티티/API)","낮음"],
 ["성능","SharedPreferences 반복 호출","매 화면마다 prefs 새로 가져옴","AuthProvider/UserProvider 상태관리 도입","중간"],
 ["성능","이미지 캐싱","Image.network 직접 사용","cached_network_image 패키지 도입","낮음"],
 ["아키텍처","상태관리 없음","setState만 사용","Provider/Riverpod 도입 검토","중간"],
]
for i,row in enumerate(opt_data,2):
    f = None
    if row[4] == "높음": f = err_fill
    elif row[4] == "중간": f = warn_fill
    add_row(ws4, i, row, f)
for col in ws4.columns:
    ws4.column_dimensions[col[0].column_letter].width = 30

path = r"c:\project\OurT\OurT_프로젝트_분석.xlsx"
wb.save(path)
print(f"OK: {path}")
