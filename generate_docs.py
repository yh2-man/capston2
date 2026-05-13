import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='2E7D32')

def style_header(cell, fgColor='2E7D32'):
    cell.font = header_font
    cell.fill = PatternFill('solid', fgColor=fgColor)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

def style_cell(cell, bold=False, color=None):
    cell.font = Font(bold=bold, size=10, color=color or '000000')
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border = border

def set_fill(cell, fgColor):
    cell.fill = PatternFill('solid', fgColor=fgColor)

# ── 시트1: Flutter-Backend API 매핑 ──────────────────────────
ws1 = wb.active
ws1.title = 'API 매핑'

h1 = ['Flutter 서비스', '메서드', 'HTTP', '백엔드 엔드포인트', '파라미터', '반환값', '상태']
for i, h in enumerate(h1, 1):
    style_header(ws1.cell(1, i, h), '1B5E20')

mappings = [
    ('CommunityApiService', 'getPosts()', 'GET', '/api/v1/community/posts', 'cursor, limit', 'List<CommunityPost>', 'OK'),
    ('CommunityApiService', 'createPost()', 'POST', '/api/v1/community/posts', 'content, species, image, lat, lng', 'CommunityPost?', 'OK'),
    ('CommunityApiService', 'toggleLike()', 'POST', '/api/v1/community/posts/{id}/like', 'postId', 'Map', 'OK'),
    ('CommunityApiService', 'toggleSave()', 'POST', '/api/v1/community/posts/{id}/save', 'postId', 'Map', 'OK'),
    ('FlowerBookApiService', 'getByMonth()', 'GET', '/api/v1/flowers/monthly/{month}', 'month(1-12)', 'List<FlowerBookItem>', 'OK'),
    ('FlowerBookApiService', 'getDetail()', 'GET', '/api/v1/flowers/{id}', 'id', 'FlowerBookDetail', 'OK'),
    ('FlowerBookApiService', 'search()', 'GET', '/api/v1/flowers/search?keyword=', 'keyword', 'List<FlowerBookItem>', 'OK'),
    ('AuthApiService', 'kakaoLogin()', 'POST', '/api/v1/auth/oauth/kakao', 'code', 'AuthResponse', 'OK'),
    ('AuthApiService', 'setupProfile()', 'POST', '/api/v1/auth/profile-setup', 'nickname, image', 'AuthResponse', 'OK'),
    ('AuthApiService', 'saveFcmToken()', 'POST', '/api/v1/auth/fcm-token', 'fcmToken', 'void', 'OK'),
    ('WalkApiService', 'getWeeklyRecords()', 'GET', '/api/v1/walk/records/weekly', 'token', 'List<WalkRecord>', 'OK'),
    ('WalkApiService', 'getPointBalance()', 'GET', '/api/v1/walk/points', 'token', 'int', 'OK'),
    ('WalkApiService', 'syncSteps()', 'POST', '/api/v1/walk/sync', 'stepCount', 'void', 'OK'),
    ('SavedApiService', 'getSavedPosts()', 'GET', '/api/v1/saved/posts', 'token', 'List<SavedPostItem>', 'OK'),
    ('SavedApiService', 'getSavedSpots()', 'GET', '/api/v1/saved/spots', 'token', 'List<SavedSpotItem>', 'OK'),
    ('SavedApiService', 'unsavePost()', 'DELETE', '/api/v1/saved/posts/{id}', 'postId', 'void', 'OK'),
    ('SavedApiService', 'unsaveSpot()', 'DELETE', '/api/v1/saved/spots/{id}', 'spotId', 'void', 'OK'),
    ('ChatbotService', 'sendMessage()', 'POST', '/chatbot/message', 'message, sessionId, lat, lng', 'ChatbotResponse', 'OK'),
    ('ChatbotService', 'clearSession()', 'DELETE', '/chatbot/session/{id}', 'sessionId', 'void', 'OK'),
    ('-', '(미연결)', 'POST', '/api/v1/admin/flowers/import', '-', '-', 'WARN: 앱 미연결'),
    ('-', '(미연결)', 'POST', '/api/v1/admin/flowers/fetch-images', '-', '-', 'WARN: 앱 미연결'),
    ('-', '(미연결)', 'GET', '/api/v1/flowers/categories', '-', '-', 'WARN: Flutter 미사용'),
    ('-', '(미연결)', 'GET', '/api/v1/flowers/match', '-', '-', 'WARN: Flutter 미사용'),
    ('map/app.js', 'fetchFlowersFromApi()', 'GET', '/flowers?lat=&lng= (없는 API)', 'lat,lng,radius', 'List', 'ERROR: 백엔드 없음'),
]

fill_colors = {
    'CommunityApiService': 'FFF8E1',
    'FlowerBookApiService': 'E8F5E9',
    'AuthApiService': 'E8EAF6',
    'WalkApiService': 'E3F2FD',
    'SavedApiService': 'E0F2F1',
    'ChatbotService': 'EDE7F6',
    '-': 'FFF3E0',
    'map/app.js': 'FFEBEE',
}

for row, data in enumerate(mappings, 2):
    fg = fill_colors.get(data[0], 'FFFFFF')
    for col, val in enumerate(data, 1):
        c = ws1.cell(row, col, val)
        style_cell(c)
        set_fill(c, fg)
        if val == 'OK':
            c.font = Font(color='1B5E20', bold=True, size=10)
            c.value = 'OK'
        elif val and 'WARN' in str(val):
            c.font = Font(color='E65100', bold=True, size=10)
        elif val and 'ERROR' in str(val):
            c.font = Font(color='B71C1C', bold=True, size=10)

for col, w in zip('ABCDEFG', [22, 25, 8, 42, 28, 18, 22]):
    ws1.column_dimensions[col].width = w

# ── 시트2: Flutter 화면 구조 ──────────────────────────────────
ws2 = wb.create_sheet('Flutter 화면 구조')
h2 = ['화면', '유형', '이름', '타입', '설명']
for i, h in enumerate(h2, 1):
    style_header(ws2.cell(1, i, h), '1A237E')

screens = [
    ('MainScreen', 'State', '_posts', 'List<CommunityPost>', '홈 게시글 미리보기 목록'),
    ('MainScreen', 'State', '_isChatRunning', 'bool', '챗봇 실행 중 여부'),
    ('MainScreen', 'State', '_chatStatus', 'String?', '챗봇 상태 메시지'),
    ('MainScreen', 'State', '_festivalIndex', 'int', '현재 배너 인덱스'),
    ('MainScreen', 'Method', '_loadPosts()', 'Future<void>', '커뮤니티 피드 로드'),
    ('MainScreen', 'Method', '_sendMainChatMessage()', 'Future<void>', '홈 챗봇 전송'),
    ('', '', '', '', ''),
    ('KakaoMapScreen', 'State', '_controller', 'WebViewController?', 'WebView 컨트롤러'),
    ('KakaoMapScreen', 'State', '_searchController', 'TextEditingController', '검색창'),
    ('KakaoMapScreen', 'State', '_isLoading', 'bool', '맵 로딩'),
    ('KakaoMapScreen', 'Public', 'setSearchQuery()', 'Future<void>', 'JS FlowerMap 브릿지'),
    ('KakaoMapScreen', 'Public', 'zoomIn/zoomOut()', 'Future<void>', 'JS FlowerMap 브릿지'),
    ('KakaoMapScreen', 'Public', 'moveToCurrentLocation()', 'Future<void>', 'Geolocator -> JS 주입'),
    ('KakaoMapScreen', 'Public', '_applyInitialActions()', 'Future<void>', '챗봇 MAP 액션 처리'),
    ('', '', '', '', ''),
    ('CommunityFeedScreen', 'State', '_posts', 'List<CommunityPost>', '게시글 목록'),
    ('CommunityFeedScreen', 'State', '_accessToken', 'String', '인증 토큰'),
    ('CommunityFeedScreen', 'State', '_error', 'String?', '에러 메시지'),
    ('CommunityFeedScreen', 'Method', '_loadPosts()', 'Future<void>', '피드 로드 + 에러처리'),
    ('CommunityFeedScreen', 'Method', '_toggleLike()', 'Future<void>', '낙관적 업데이트+롤백'),
    ('CommunityFeedScreen', 'Method', '_toggleSave()', 'Future<void>', '낙관적 업데이트+롤백'),
    ('', '', '', '', ''),
    ('FlowerBookPage', 'State', '_flowers', 'List<FlowerBookItem>', '꽃 목록'),
    ('FlowerBookPage', 'State', '_selectedMonth', 'int', '선택된 월 (1-12)'),
    ('FlowerBookPage', 'State', '_isSearching', 'bool', '검색 모드'),
    ('FlowerBookPage', 'State', '_error', 'String?', '에러'),
    ('FlowerBookPage', 'Method', '_loadFlowers()', 'Future<void>', '월별 꽃 로드'),
    ('FlowerBookPage', 'Method', '_search()', 'Future<void>', '키워드 검색'),
    ('', '', '', '', ''),
    ('PedometerScreen', 'State', '_steps', 'int', '현재 걸음수'),
    ('PedometerScreen', 'State', '_pointBalance', 'int', '포인트 잔액'),
    ('PedometerScreen', 'State', '_weeklyData', 'List<WalkRecord>', '주간 기록'),
    ('PedometerScreen', 'State', '_isLoading', 'bool', '로딩'),
    ('PedometerScreen', 'Method', '_loadData()', 'Future<void>', 'API 로드 + try-catch'),
    ('PedometerScreen', 'Method', '_simulateWalk()', 'Future<void>', '+150보 + 동기화'),
    ('', '', '', '', ''),
    ('SavedPage', 'State', '_savedPosts', 'List<SavedPostItem>', '저장된 게시글'),
    ('SavedPage', 'State', '_savedSpots', 'List<SavedSpotItem>', '저장된 명소'),
    ('SavedPage', 'State', '_tabController', 'TabController', '탭 컨트롤러'),
    ('SavedPage', 'Method', 'onDismissed()', 'Future<void>', '삭제 + 실패시 rollback'),
]

screen_fills = {
    'MainScreen': 'E3F2FD',
    'KakaoMapScreen': 'E8F5E9',
    'CommunityFeedScreen': 'FFF3E0',
    'FlowerBookPage': 'F3E5F5',
    'PedometerScreen': 'FCE4EC',
    'SavedPage': 'E0F2F1',
}

cur = ''
for row, data in enumerate(screens, 2):
    if data[0]: cur = data[0]
    fg = screen_fills.get(cur, 'FFFFFF')
    for col, val in enumerate(data, 1):
        c = ws2.cell(row, col, val)
        style_cell(c, bold=(col == 1 and bool(val)))
        if val: set_fill(c, fg)

for col, w in zip('ABCDE', [20, 10, 28, 22, 35]):
    ws2.column_dimensions[col].width = w

# ── 시트3: 백엔드 API 전체 ──────────────────────────────────────
ws3 = wb.create_sheet('백엔드 API 목록')
h3 = ['컨트롤러', 'HTTP', '엔드포인트', '인증', '설명', 'Flutter 연결']
for i, h in enumerate(h3, 1):
    style_header(ws3.cell(1, i, h), '880E4F')

endpoints = [
    ('AuthController', 'POST', '/api/v1/auth/oauth/kakao', 'X', '카카오 OAuth 로그인', 'OK'),
    ('AuthController', 'POST', '/api/v1/auth/refresh', 'X', '토큰 갱신', 'OK'),
    ('AuthController', 'POST', '/api/v1/auth/profile-setup', 'X', '신규 사용자 프로필', 'OK'),
    ('AuthController', 'POST', '/api/v1/auth/fcm-token', 'O', 'FCM 토큰 저장', 'OK'),
    ('CommunityController', 'GET', '/api/v1/community/posts', 'O', '게시글 피드 (커서 페이지네이션)', 'OK'),
    ('CommunityController', 'POST', '/api/v1/community/posts', 'O', '게시글 작성 (multipart)', 'OK'),
    ('CommunityController', 'POST', '/api/v1/community/posts/{id}/like', 'O', '좋아요 토글', 'OK'),
    ('CommunityController', 'POST', '/api/v1/community/posts/{id}/save', 'O', '저장 토글', 'OK'),
    ('FlowerController', 'GET', '/api/v1/flowers/categories', 'X', '카테고리 목록', 'WARN: Flutter 미사용'),
    ('FlowerController', 'GET', '/api/v1/flowers/categories/{id}', 'X', '카테고리별 꽃 목록', 'WARN: Flutter 미사용'),
    ('FlowerController', 'GET', '/api/v1/flowers/monthly/{month}', 'X', '월별 꽃 목록', 'OK'),
    ('FlowerController', 'GET', '/api/v1/flowers/{id}', 'X', '꽃 상세 정보', 'OK'),
    ('FlowerController', 'GET', '/api/v1/flowers/search?keyword=', 'X', '꽃 이름 검색', 'OK'),
    ('FlowerController', 'GET', '/api/v1/flowers/match?scientificName=', 'X', 'Plant.id 학명 매칭', 'WARN: Flutter 미사용'),
    ('FlowerImportController', 'POST', '/api/v1/admin/flowers/import', 'X', '농사로 데이터 수집 (1회성)', 'WARN: 앱 미연결'),
    ('FlowerImportController', 'POST', '/api/v1/admin/flowers/fetch-images', 'X', '위키 이미지 수집', 'WARN: 앱 미연결'),
    ('FlowerImportController', 'POST', '/api/v1/admin/flowers/compress-images', 'X', '농사로 이미지 압축', 'WARN: 앱 미연결'),
    ('ChatbotController', 'POST', '/chatbot/message', 'X', 'AI 챗봇 메시지', 'OK'),
    ('ChatbotController', 'DELETE', '/chatbot/session/{id}', 'X', '챗봇 세션 초기화', 'OK'),
    ('WalkController', 'GET', '/api/v1/walk/records/weekly', 'O', '주간 걸음 기록', 'OK'),
    ('WalkController', 'GET', '/api/v1/walk/points', 'O', '포인트 잔액', 'OK'),
    ('WalkController', 'POST', '/api/v1/walk/sync', 'O', '걸음수 동기화', 'OK'),
    ('SavedController', 'GET', '/api/v1/saved/posts', 'O', '저장된 게시글 목록', 'OK'),
    ('SavedController', 'GET', '/api/v1/saved/spots', 'O', '저장된 명소 목록', 'OK'),
    ('SavedController', 'DELETE', '/api/v1/saved/posts/{id}', 'O', '게시글 저장 취소', 'OK'),
    ('SavedController', 'DELETE', '/api/v1/saved/spots/{id}', 'O', '명소 저장 취소', 'OK'),
    ('ERROR: 미구현', 'GET', '/flowers?lat=&lng=&radius=', '-', '지도용 꽃 명소 (app.js 호출)', 'ERROR: 404 발생'),
]

ctrl_fills = {
    'AuthController': 'E8EAF6',
    'CommunityController': 'FFF8E1',
    'FlowerController': 'E8F5E9',
    'FlowerImportController': 'FBE9E7',
    'ChatbotController': 'EDE7F6',
    'WalkController': 'E3F2FD',
    'SavedController': 'E0F2F1',
    'ERROR: 미구현': 'FFEBEE',
}

for row, data in enumerate(endpoints, 2):
    fg = ctrl_fills.get(data[0], 'FFFFFF')
    for col, val in enumerate(data, 1):
        c = ws3.cell(row, col, val)
        style_cell(c)
        set_fill(c, fg)
        if val == 'OK':
            c.font = Font(color='1B5E20', bold=True, size=10)
        elif val and 'WARN' in str(val):
            c.font = Font(color='E65100', bold=True, size=10)
        elif val and 'ERROR' in str(val):
            c.font = Font(color='B71C1C', bold=True, size=10)

for col, w in zip('ABCDEF', [22, 8, 42, 8, 30, 22]):
    ws3.column_dimensions[col].width = w

# 행 높이
for ws in [ws1, ws2, ws3]:
    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

wb.save('c:/project/OurT/OurT_API_문서.xlsx')
print('완료: c:/project/OurT/OurT_API_문서.xlsx')
